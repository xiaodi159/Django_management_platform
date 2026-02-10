import copy

from django.shortcuts import render, HttpResponse, redirect
from stalfApp import models
from openpyxl import load_workbook

from stalfApp.models import Department
from django.utils.safestring import mark_safe
from stalfApp.utils.pagination import Pagination


#from django.http.request import QueryDict


# Create your views here.

def depart_list(request):
    data_list = Department.objects.all()
    """部门列表"""
    return render(
        request,
        'depart_list.html',
        {'data_list':data_list}
    )

#添加部门列表
def depart_add(request):
    if(request.method == "GET"):
        return render(request, 'depart_add.html')

    #获取数据
    title = request.POST.get("title")

    #保存到数据库
    Department.objects.create(title=title)
    #重定向回部门列表
    return redirect("/depart/list/")


#删除部门
def depart_delete(request):
    nid = request.GET.get("nid")

    Department.objects.filter(id=nid).delete()
    return redirect("/depart/list/")

#修改信息
def depart_edit(request, nid):
    if(request.method == "GET"):
        row_object = Department.objects.filter(id=nid).first()
        return render(request, 'depart_edit.html',{'row_object':row_object})

    new_name = request.POST.get("new_title")
    Department.objects.filter(id=nid).update(title=new_name)
    return redirect("/depart/list")

def depart_multi(request):
    """通过Excel表格上传数据"""
    # 获取用户上传的文件对象
    file_object = request.FILES.get('exc')
    print(file_object.name)
    # print(file_object.content_type)
    # print(type(file_object))

    """
        python文档操作
        from openpyxl import load_workbook
        wb = load_workbook("文件路径") or wb = load_workbook(文件对象)
        # 获取表格内容
        sheet = wb.worksheets[0]
    """

    # 通过文件对象上传至openpyxl
    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]
    # print(sheet.cell(1, 1).value)

    # 循环获取所有数据
    for row in sheet.iter_rows(min_row=1):
        text = row[0].value
        print(text)

        row_object = Department.objects.filter(title=text).first()
        if not row_object:
            Department.objects.create(title=text)

    return redirect("/depart/list/")

